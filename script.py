import argparse
import requests
import json
import openpyxl
from openpyxl.styles import Alignment

# Replace with your Meraki API key
MERAKI_API_KEY = '265ec40a39f861984f644977ca0c4d8abf586234'

def get_organization_id(org_name):
    """
    Retrieves the organization ID for the given organization name.
    """
    url = f'https://api.meraki.com/api/v1/organizations'
    headers = {
        'X-Cisco-Meraki-API-Key': MERAKI_API_KEY,
        'Content-Type': 'application/json'
    }
    response = requests.get(url, headers=headers)
    organizations = response.json()

    for org in organizations:
        if org['name'] == org_name:
            return org['id']
    return None

def get_network_topology(network_id):
    """
    Retrieves the network topology for the given network ID.
    """
    url = f'https://api.meraki.com/api/v1/networks/{network_id}/topology/list'
    headers = {
        'X-Cisco-Meraki-API-Key': MERAKI_API_KEY,
        'Content-Type': 'application/json'
    }
    response = requests.get(url, headers=headers)
    return response.json()

def get_wan_stats(network_id):
    """
    Retrieves the WAN statistics for the given network ID.
    """
    url = f'https://api.meraki.com/api/v1/networks/{network_id}/uplinks'
    headers = {
        'X-Cisco-Meraki-API-Key': MERAKI_API_KEY,
        'Content-Type': 'application/json'
    }
    response = requests.get(url, headers=headers)
    return response.json()

def get_switch_stacks(network_id):
    """
    Retrieves the switch stack information for the given network ID.
    """
    url = f'https://api.meraki.com/api/v1/networks/{network_id}/switchStacks'
    headers = {
        'X-Cisco-Meraki-API-Key': MERAKI_API_KEY,
        'Content-Type': 'application/json'
    }
    response = requests.get(url, headers=headers)
    return response.json()

def write_to_excel(org_name, network_name, topology, wan_stats, switch_stacks):
    """
    Writes the Meraki topology, WAN statistics, and switch stack information to an Excel spreadsheet.
    """
    workbook = openpyxl.Workbook()

    # Topology sheet
    topology_sheet = workbook.active
    topology_sheet.title = 'Topology'
    topology_sheet['A1'] = 'Organization'
    topology_sheet['B1'] = 'Network'
    topology_sheet['C1'] = 'Node Name'
    topology_sheet['D1'] = 'Node Type'
    topology_sheet['E1'] = 'Uplink Interface'
    topology_sheet['F1'] = 'Uplink Name'
    topology_sheet['G1'] = 'Uplink IP'
    topology_sheet['H1'] = 'Uplink MAC'

    row = 2
    for node in topology:
        topology_sheet.cell(row=row, column=1, value=org_name)
        topology_sheet.cell(row=row, column=2, value=network_name)
        topology_sheet.cell(row=row, column=3, value=node['name'])
        topology_sheet.cell(row=row, column=4, value=node['type'])
        topology_sheet.cell(row=row, column=5, value=node['uplink']['interface'])
        topology_sheet.cell(row=row, column=6, value=node['uplink']['name'])
        topology_sheet.cell(row=row, column=7, value=node['uplink']['ip'])
        topology_sheet.cell(row=row, column=8, value=node['uplink']['mac'])
        row += 1

    # WAN stats sheet
    wan_stats_sheet = workbook.create_sheet('WAN Statistics')
    wan_stats_sheet['A1'] = 'Organization'
    wan_stats_sheet['B1'] = 'Network'
    wan_stats_sheet['C1'] = 'Uplink Name'
    wan_stats_sheet['D1'] = 'Uplink Interface'
    wan_stats_sheet['E1'] = 'Public IP'
    wan_stats_sheet['F1'] = 'Public Hostname'
    wan_stats_sheet['G1'] = 'Public Latency'
    wan_stats_sheet['H1'] = 'Public Packet Loss'
    wan_stats_sheet['I1'] = 'Public Bandwidth'

    row = 2
    for uplink in wan_stats:
        wan_stats_sheet.cell(row=row, column=1, value=org_name)
        wan_stats_sheet.cell(row=row, column=2, value=network_name)
        wan_stats_sheet.cell(row=row, column=3, value=uplink['name'])
        wan_stats_sheet.cell(row=row, column=4, value=uplink['interface'])
        wan_stats_sheet.cell(row=row, column=5, value=uplink['publicIp'])
        wan_stats_sheet.cell(row=row, column=6, value=uplink['publicHostname'])
        wan_stats_sheet.cell(row=row, column=7, value=uplink['publicLatency'])
        wan_stats_sheet.cell(row=row, column=8, value=uplink['publicPacketLoss'])
        wan_stats_sheet.cell(row=row, column=9, value=uplink['publicBandwidth'])
        row += 1

    # Switch stack sheet
    switch_stack_sheet = workbook.create_sheet('Switch Stacks')
    switch_stack_sheet['A1'] = 'Organization'
    switch_stack_sheet['B1'] = 'Network'
    switch_stack_sheet['C1'] = 'Switch Stack Name'
    switch_stack_sheet['D1'] = 'Switch Model'
    switch_stack_sheet['E1'] = 'Switch Serial'
    switch_stack_sheet['F1'] = 'Switch Status'

    row = 2
    for stack in switch_stacks:
        for member in stack['members']:
            switch_stack_sheet.cell(row=row, column=1, value=org_name)
            switch_stack_sheet.cell(row=row, column=2, value=network_name)
            switch_stack_sheet.cell(row=row, column=3, value=stack['name'])
            switch_stack_sheet.cell(row=row, column=4, value=member['model'])
            switch_stack_sheet.cell(row=row, column=5, value=member['serial'])
            switch_stack_sheet.cell(row=row, column=6, value=member['status'])
            row += 1

    # Adjust column widths
    for column in range(1, 10):
        topology_sheet.column_dimensions[chr(column + 64)].auto_size = True
        wan_stats_sheet.column_dimensions[chr(column + 64)].auto_size = True
        switch_stack_sheet.column_dimensions[chr(column + 64)].auto_size = True

    # Center align the headers
    for row in range(1, 2):
        for col in range(1, 10):
            cell = topology_sheet.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center')
            cell = wan_stats_sheet.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center')
            cell = switch_stack_sheet.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center')

    # Save the Excel file
    workbook.save(f'{org_name}_{network_name}.xlsx')
    print(f'Excel file saved: {org_name}_{network_name}.xlsx')

def main():
    # Create the argument parser
    parser = argparse.ArgumentParser(description='Meraki Topology, WAN Statistics, and Switch Stack Information')
    parser.add_argument('--org-name', type=str, required=True, help='The name of the Meraki organization')
    parser.add_argument('--network-name', type=str, required=True, help='The name of the Meraki network')

    # Parse the arguments
    args = parser.parse_args()

    # Get the organization ID
    org_id = get_organization_id(args.org_name)
    if not org_id:
        print(f'Error: Organization "{args.org_name}" not found.')
        return

    # Get the network ID
    url = f'https://api.meraki.com/api/v1/organizations/{org_id}/networks'
    headers = {
        'X-Cisco-Meraki-API-Key': MERAKI_API_KEY,
        'Content-Type': 'application/json'
    }
    response = requests.get(url, headers=headers)
    networks = response.json()

    network_id = None
    for network in networks:
        if network['name'] == args.network_name:
            network_id = network['id']
            break

    if not network_id:
        print(f'Error: Network "{args.network_name}" not found.')
        return

    # Get the topology, WAN statistics, and switch stack information
    topology = get_network_topology(network_id)
    wan_stats = get_wan_stats(network_id)
    switch_stacks = get_switch_stacks(network_id)

    # Write the results to an Excel spreadsheet
    write_to_excel(args.org_name, args.network_name, topology, wan_stats, switch_stacks)

if __name__ == '__main__':
    main()